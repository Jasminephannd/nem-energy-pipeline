"""Build dim_unit from AEMO's NEM Registration and Exemption List.

The registration list is the authoritative DUID -> station / region / fuel /
technology / capacity mapping. Its "PU and Scheduled Loads" sheet is what
gives fct_unit_dispatch meaning: without a fuel type there is no
renewable-penetration analysis.

Classification uses AEMO's own STRUCTURED columns rather than keyword
guessing (confirmed by inspecting the real workbook):

  * Technology Type - Primary : Renewable | Combustion | Storage
  * Fuel Source - Primary     : Solar | Wind | Hydro | Fossil | Battery
                                Storage | Renewable/ Biomass / Waste
  * Dispatch Type             : Generating Unit | Bidirectional Unit | Load

is_renewable  = Technology Primary is 'Renewable', OR Fuel Primary is a
                renewable category. Batteries are NOT renewable generation
                (storage returns grid energy), but get is_storage = 1 so they
                are a distinct dashboard category rather than lumped with fossil.

Run (venv active):
    python scripts/build_dim_unit.py                       # download + inspect
    python scripts/build_dim_unit.py --emit-sql sql/04_dim_unit.sql
    python scripts/build_dim_unit.py --xlsx local.xlsx     # use a file you downloaded
"""

from __future__ import annotations

import argparse
import io
import sys
from collections import Counter

import openpyxl
import requests

AEMO_URL = (
    "https://www.aemo.com.au/-/media/files/electricity/nem/"
    "participant_information/nem-registration-and-exemption-list.xlsx"
)
_HEADERS = {"User-Agent": "nem-energy-pipeline/1.0 (portfolio project)"}

# Column headers we need, matched case-insensitively against the header row.
COLUMN_HINTS = {
    "station": ["station name"],
    "region": ["region"],
    "fuel_primary": ["fuel source - primary"],
    "fuel_detail": ["fuel source - descriptor"],
    "tech_primary": ["technology type - primary"],
    "duid": ["duid"],
    "capacity": ["reg cap generation"],
}


def energy_category(tech_primary: str) -> str:
    """Map AEMO's Technology Type - Primary to a clean 3-way category.

    A single mutually-exclusive category (Renewable | Combustion | Storage)
    rather than overlapping boolean flags: Combustion is a real value, not the
    absence of the others, and there are no invalid states. is_renewable is
    then just a convenience flag derived from this.
    """
    t = tech_primary.strip().lower()
    if t == "renewable":
        return "Renewable"
    if t == "combustion":
        return "Combustion"
    if t == "storage":
        return "Storage"
    return "Unknown"  # the handful of '-' / blank rows


def load_workbook_bytes(xlsx_path: str | None) -> bytes:
    if xlsx_path:
        with open(xlsx_path, "rb") as fh:
            return fh.read()
    print(f"downloading {AEMO_URL} ...")
    resp = requests.get(AEMO_URL, headers=_HEADERS, timeout=120)
    resp.raise_for_status()
    print(f"  {len(resp.content):,} bytes")
    return resp.content


def find_sheet(wb) -> str:
    # "PU and Scheduled Loads" (formerly "Generators and Scheduled Loads").
    for name in wb.sheetnames:
        low = name.lower()
        if "scheduled" in low and "load" in low:
            return name
    for name in wb.sheetnames:
        low = name.lower()
        if "generator" in low or low.startswith("pu"):
            return name
    raise SystemExit(f"No PU/generators sheet found. Sheets: {wb.sheetnames}")


def find_header_row(ws, max_scan: int = 15) -> int:
    for r in range(1, max_scan + 1):
        values = [str(c.value).lower() if c.value else "" for c in ws[r]]
        if any("duid" in v for v in values):
            return r
    raise SystemExit("Could not find a header row containing 'DUID'.")


def map_columns(ws, header_row: int) -> dict[str, int]:
    headers = {i: (str(c.value).strip().lower() if c.value else "")
               for i, c in enumerate(ws[header_row])}
    found: dict[str, int] = {}
    for key, hints in COLUMN_HINTS.items():
        for hint in hints:
            match = next((i for i, h in headers.items() if hint in h), None)
            if match is not None:
                found[key] = match
                break
    return found


def to_region_id(region: str) -> str:
    r = region.strip().upper()
    mapping = {"NSW": "NSW1", "QLD": "QLD1", "VIC": "VIC1", "SA": "SA1", "TAS": "TAS1"}
    return mapping.get(r, r)  # AEMO already uses NSW1 etc.; pass through


def extract_rows(ws, header_row: int, cols: dict[str, int]) -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()
    for r in range(header_row + 1, ws.max_row + 1):
        cells = ws[r]

        def get(key: str) -> str:
            i = cols.get(key)
            v = cells[i].value if i is not None else None
            return str(v).strip() if v is not None else ""

        duid = get("duid").upper()
        if not duid or duid in seen:
            continue
        seen.add(duid)

        category = energy_category(get("tech_primary"))
        rows.append({
            "duid": duid,
            "station": get("station"),
            "region_id": to_region_id(get("region")),
            "energy_category": category,
            "fuel_category": get("fuel_primary"),
            "fuel_detail": get("fuel_detail"),
            "is_renewable": int(category == "Renewable"),
            "capacity": get("capacity"),
        })
    return rows


def _sql(s: str) -> str:
    return s.replace("'", "''")


def _num(s: str) -> str:
    try:
        return str(float(s.replace(",", "")))
    except ValueError:
        return "NULL"


def emit_sql(rows: list[dict], path: str) -> None:
    lines = [
        "/* dim_unit - generated from AEMO NEM Registration and Exemption List",
        "   ('PU and Scheduled Loads' sheet) by scripts/build_dim_unit.py.",
        "   Grain: one row per generating/storage unit (DUID).",
        "   energy_category is a direct copy of AEMO's Technology Type - Primary",
        "   (Renewable | Combustion | Storage); is_renewable is derived from it",
        "   as a convenience for the headline renewable-% measure.",
        "   No FK from fct_unit_dispatch on purpose: some SCADA DUIDs (retired",
        "   or renamed) may not be in the current list - a LEFT JOIN reports",
        "   those as unclassified rather than rejecting them. */",
        "IF OBJECT_ID('dbo.dim_unit','U') IS NOT NULL DROP TABLE dbo.dim_unit;",
        "CREATE TABLE dbo.dim_unit (",
        "    duid            VARCHAR(20)  NOT NULL CONSTRAINT pk_dim_unit PRIMARY KEY,",
        "    station_name    VARCHAR(100) NULL,",
        "    region_id       VARCHAR(10)  NULL,",
        "    energy_category VARCHAR(12)  NOT NULL,  -- Renewable | Combustion | Storage | Unknown",
        "    fuel_category   VARCHAR(40)  NULL,      -- Solar | Wind | Hydro | Fossil | Battery Storage | ...",
        "    fuel_detail     VARCHAR(60)  NULL,      -- Black Coal | Natural Gas | Water | Diesel | ...",
        "    is_renewable    BIT          NOT NULL,  -- derived: energy_category = 'Renewable'",
        "    capacity_mw     DECIMAL(10,3) NULL",
        ");",
    ]
    for row in rows:
        lines.append(
            "INSERT INTO dbo.dim_unit (duid, station_name, region_id, energy_category, "
            "fuel_category, fuel_detail, is_renewable, capacity_mw) VALUES ("
            f"'{_sql(row['duid'])}', "
            f"'{_sql(row['station'])}', "
            f"'{_sql(row['region_id'])}', "
            f"'{_sql(row['energy_category'])}', "
            f"'{_sql(row['fuel_category'])}', "
            f"'{_sql(row['fuel_detail'])}', "
            f"{row['is_renewable']}, {_num(row['capacity'])});"
        )
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    print(f"\nwrote {len(rows)} units -> {path}")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Build dim_unit from the AEMO registration list.")
    p.add_argument("--xlsx", help="Local .xlsx instead of downloading.")
    p.add_argument("--emit-sql", help="Write CREATE + INSERTs to this .sql path.")
    p.add_argument("--sample", type=int, default=12, help="Rows to preview.")
    args = p.parse_args(argv if argv is not None else sys.argv[1:])

    wb = openpyxl.load_workbook(io.BytesIO(load_workbook_bytes(args.xlsx)),
                                read_only=True, data_only=True)
    sheet = find_sheet(wb)
    ws = wb[sheet]
    header_row = find_header_row(ws)
    cols = map_columns(ws, header_row)

    print(f"\nsheet: {sheet!r}   header row: {header_row}")
    print("detected columns:")
    for key in COLUMN_HINTS:
        print(f"  {key:14} -> column index {cols.get(key, 'NOT FOUND')}")
    missing = [k for k in ("duid", "tech_primary", "fuel_primary") if k not in cols]
    if missing:
        raise SystemExit(f"Missing essential columns {missing} - inspect manually.")

    rows = extract_rows(ws, header_row, cols)
    ren = sum(r["is_renewable"] for r in rows)
    print(f"\n{len(rows)} units | renewable {ren} | non-renewable {len(rows) - ren}")
    print("energy_category:",
          dict(Counter(r["energy_category"] for r in rows).most_common()))
    print("fuel_category:  ",
          dict(Counter(r["fuel_category"] for r in rows).most_common()))
    print(f"\n--- first {args.sample} ---")
    for r in rows[:args.sample]:
        print(f"  {r['duid']:12} {r['energy_category']:10} ren={r['is_renewable']} "
              f"{r['fuel_category']:16} {r['region_id']:5} {r['station'][:26]}")

    if args.emit_sql:
        emit_sql(rows, args.emit_sql)
    else:
        print("\n(looks right? re-run with --emit-sql sql/04_dim_unit.sql)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
